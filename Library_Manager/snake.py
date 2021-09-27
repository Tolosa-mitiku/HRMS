from tkinter import *
import tkinter.messagebox
from tkinter import ttk
import time
import datetime
import os
import openpyxl

def main():
    root = Tk()
    Window2(root)                                                                                                #calling the Window2 class inseting the root as a parameter

class Student:
    def __init__(self,win):
        self.win =win                                                                                            #window
        self.win.title("Library Management system")
        self.win.resizable(width =False, height=False)                                                           # to make the window size static
        self.win.config(bg = 'dark grey')
        self.frame = Frame(self.win,bg='dark grey')
        self.frame.pack()

        self.Username = StringVar()
        self.Password = StringVar()

        self.lblTitle  = Label(self.frame, text = "Library Log in Checker",font =("Algerian",20,'bold'),bg = 'dark grey',fg = 'black')
        self.lblTitle.grid(row = 0,column=0, columnspan = 2,pady = 20)


#++++++++++++++++++++++++++++++++++++Lable and entry++++++++++++++++++++++++++++++++++++++++++++++++++++

        self.lblUsername = Label(self.frame,text='Username',font=('arial',12,'bold'),bg="dark grey")
        self.lblUsername.grid(row=3,column=0)
        self.txtUsername = Entry(self.frame,text='Username',font=('arial',12,'bold'),textvariable=self.Username)
        self.txtUsername.grid(row = 3,column =1)

        self.lblPassword = Label(self.frame,text='Password',font=('arial',12,'bold'),bg="dark grey")
        self.lblPassword.grid(row=4,column=0)
        self.txtPassword = Entry(self.frame,text='Password',font=('arial',12,'bold'),show="●",textvariable= self.Password)
        self.txtPassword.grid(row = 4,column =1)
#++++++++++++++++++++++++++++++++++++++++++++++++++++++Button++++++++++++++++++++++++++++++++++    
        self.btnLogin = Button(self.frame, text = 'Login', width = 17,font=('arial',12,'bold'), command= self.Login_System)
        self.btnLogin.grid(row=5,column = 0,pady=20,padx =8)

        self.btnExit = Button(self.frame, text = 'Exit', width = 17,font=('arial',12,'bold'), command= self.iExit)
        self.btnExit.grid(row=5,column = 1,pady=20,padx =8)

        self.btnReset = Button(self.frame, text = 'Reset', width = 17,font=('arial',12,'bold'),command= self.reset)
        self.btnReset.grid(row=5,column = 2,pady=20,padx =8)

#++++++++++++++++++++++++++++++++Button++++++++++++++++++++++++++++++++++
    def Login_System(self):
      
        u = (self.Username.get())
        p = (self.Password.get())
        if u == str('S') and p == str('P'):                                                                            #the password and username must be 'S' and 'P' in order to log in.
             self.newWindow=Toplevel(self.win)
             self.app = Book_Tracker_Stud(self.newWindow)
            
            
        else:
            tkinter.messagebox.askyesno("Login system","Too Bad,Invalid Login detail")
            self.Username.set("")
            self.Password.set("")
            self.txtUsername.focus()
    def reset(self):
        self.Username.set("")
        self.Password.set("")
        self.txtUsername.focus()
    def iExit(self):
        self.iExit=tkinter.messagebox.askyesno("Login system","Confirm to exit")
        if self.iExit>0:
            self.win.destroy()
        else:
            command=self.new_window
            return
            




    def new_window(self):
        self.newWindow=Toplevel(self.win)
        self.app = Window2(self.newWindow)

#++++++++++++++++++++++++++++++++++++++++++++++ class end+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
class Teacher:
    def __init__(self,win):
        self.win =win                                                                         #window
        self.win.title("Library Management system")
        self.win.resizable(width =False, height=False)                                        # to make the window size static
        self.win.config(bg = 'dark grey')
        self.frame = Frame(self.win,bg='dark grey')
        self.frame.pack()

        self.Username = StringVar()
        self.Password = StringVar()

        self.lblTitle  = Label(self.frame, text = "Library Log in Checker",font =("Algerian",20,'bold'),bg = 'dark grey',fg = 'black')
        self.lblTitle.grid(row = 0,column=0, columnspan = 2,pady = 20)


#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++ Lable and entry ++++++++++++++++++++++++++++++++++++++++++++++++++++

        self.lblUsername = Label(self.frame,text='Username',font=('arial',12,'bold'),bg="dark grey")
        self.lblUsername.grid(row=3,column=0)
        self.txtUsername = Entry(self.frame,text='Username',font=('arial',12,'bold'),textvariable=self.Username)
        self.txtUsername.grid(row = 3,column =1)

        self.lblPassword = Label(self.frame,text='Password',font=('arial',12,'bold'),bg="dark grey")
        self.lblPassword.grid(row=4,column=0)
        self.txtPassword = Entry(self.frame,text='Password',font=('arial',12,'bold'),show="●",textvariable= self.Password)      #hide the password
        self.txtPassword.grid(row = 4,column =1)
#++++++++++++++++++++++++++++++++++++++++++++++++++++++ Button ++++++++++++++++++++++++++++++++++    
        self.btnLogin = Button(self.frame, text = 'Login', width = 17,font=('arial',12,'bold'), command= self.Login_System)
        self.btnLogin.grid(row=5,column = 0,pady=20,padx =8)

        self.btnExit = Button(self.frame, text = 'Exit', width = 17,font=('arial',12,'bold'), command= self.iExit)
        self.btnExit.grid(row=5,column = 1,pady=20,padx =8)

        self.btnReset = Button(self.frame, text = 'Reset', width = 17,font=('arial',12,'bold'),command= self.reset)
        self.btnReset.grid(row=5,column = 2,pady=20,padx =8)
#++++++++++++++++++++++++++++++++++++++++++++++++++++++ Button ++++++++++++++++++++++++++++++++++
    def Login_System(self):
        u = (self.Username.get())                                                                              # username

        p = (self.Password.get())                                                                              # password
        
        if u == str('T') and p == str('P'):
             self.newWindow=Toplevel(self.win)
             self.app = Book_Tracker_admin(self.newWindow)
          
             
        else:
            tkinter.messagebox.askyesno("Login system","Too Bad,Invalid Login detail")
            self.Username.set("")
            self.Password.set("")
            self.txtUsername.focus()
    def reset(self):
        self.Username.set("")
        self.Password.set("")
        self.txtUsername.focus()
    def iExit(self):
        self.iExit=tkinter.messagebox.askyesno("Login system","Confirm to exit")
        if self.iExit>0:
            self.win.destroy()
        else:
            command=self.new_window
            return
        
    def new_window(self):
        self.newWindow=Toplevel(self.win)
        self.app = Window2(self.newWindow)
#++++++++++++++++++++++++++++++++class end++++++++++++++++++++++++++++++++++++++++
class Window2:                                                                                                   # log in window
    def __init__(self,win):
        self.win =win
        self.win.title("Library Management System")
  
        self.win.resizable(width =False, height=False)                                                           # to make the window size static
        self.win.config(bg = 'dark grey')
        self.frame = Frame(self.win,bg='dark grey')
        self.frame.pack()
        self.lblTitle  = Label(self.frame, text = "Wellcome Select your profile",font =("Algerian",20,'bold'),bg = 'dark grey',fg = 'black')
        self.lblTitle.grid(row = 0,column=1, columnspan = 3,pady = 40)
        
#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

        self.btnStudent = Button(self.frame, text = 'User', width = 17,font=('arial',12,'bold'), command= self.Student)
        self.btnStudent.grid(row=5,column = 1,pady=20,padx =8)

        self.btnTeacher = Button(self.frame, text = 'Librarian', width = 17,font=('arial',12,'bold'), command= self.Teacher)
        self.btnTeacher.grid(row=5,column = 3,pady=20,padx =8)
        
        self.btnExit = Button(self.frame, text = 'Exit', width = 17,font=('arial',12,'bold'), command= self.iExit)
        self.btnExit.grid(row=6,column = 2,pady=5,padx =8)
    def Student(self):
        self.newWindow=Toplevel(self.win)
        self.app = Student(self.newWindow)
   
    
    def Teacher(self):
        self.newWindow=Toplevel(self.win)
        self.app = Teacher(self.newWindow)
     
    def iExit(self):
        self.iExit=tkinter.messagebox.askyesno("Login system","Confirm to exit")
        if self.iExit>0:
            self.win.destroy()
        
#========================================student one profile======================================================

class Book_Tracker_admin: 
       
     def __init__(self,root):
          self.root = root
          self.root.title("Library Management system")
          self.root.geometry("1350x750+0+0")
          self.root.configure(background='dark grey')
          
          Ref= StringVar()
          B_title = StringVar()
          B_type = StringVar()
          Author = StringVar()
          DateBorrowed = StringVar()
          DateDue = StringVar()
          SellPrice = StringVar()
          Language=StringVar()
          Page = StringVar()
          Year=StringVar()
          Psize=StringVar()
          
          os.chdir('C:\\Users\\OKSAS\\Desktop')                                                              #importing the excell file
          File=openpyxl.load_workbook('managerdemo.xlsx')                                                    #excell file database
          sheet=File['Sheet1']                                                                               #importing sheet1 from the workbook
          
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++line break+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
          
          def Exit():
               Exit=tkinter.messagebox.askyesno("Library management system","Do you want to Log out ?")
               if Exit >0:
                    root.destroy()
                    return
        
          def Add_Book():                                                                                                #displaying the details of the selected book
              listInfo=["",int(Ref.get()),B_title.get(),Author.get(),int(Page.get()),Year.get(),Language.get(),Psize.get(),B_type.get(),SellPrice.get()]
              sheet2=File['Sheet1']                                                                                      #importing sheet2 from the workbook
              maxx_row=sheet2.max_row
              maxx_column= sheet2.max_column
              for j in range(1,len(listInfo)):
                  sheet2.cell(row=maxx_row +1,column=j).value= listInfo[j]
                  File.save('managerdemo.xlsx')
                  
#++++++++++++++++++++++++++++++++++++++++++++++++++++++++deleting a book++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

                          

          def Detail():                                                                                                #displaying the details of the selected book 
              self.txtDisplay2.config(state=NORMAL)                                                                     
              
              self.txtDisplay2.delete("1.0",END)                                                                       #clearing the text entries 
              self.txtDisplay2.insert(END,'Title: \t\t' + B_title.get() + "\n")
              self.txtDisplay2.insert(END,'Author: \t\t' + Author.get() + "\n" )
              self.txtDisplay2.insert(END,'No of page: \t\t' + Page.get() + "\n" )
                  
              
              self.txtDisplay2.insert(END,'Language :\t\t' + Language.get()+ "\n" )
              self.txtDisplay2.insert(END,'Size: \t\t' + Psize.get() + "\n" )
              self.txtDisplay2.insert(END,'Year: \t\t' + Year.get() + "\n" )
              self.txtDisplay2.insert(END,'Book ID: \t\t' + Ref.get() + "\n" )
              self.txtDisplay2.config(state=DISABLED)                                                                 #Preventing  data temper
  
              
#+++++++++++++++++++++++++++++++++++++++++++++++++++++++=reseting the book+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

              
              
              
              
          def Reset():
               Ref.set("")
               B_type.set("")
               B_title.set("")
               Author.set("")
               SellPrice.set("")
               Year.set("")
               Language.set("")
               Page.set("")
               Psize.set("")
               
              
#+++++++++++++++++++++++++++++++++++++++++++++++++++++++main frame+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
          MainFrame = Frame(self.root)
          MainFrame.grid()

          TitleFrame = Frame(MainFrame,width = 1350,bg="dark grey",padx = 20,bd =20,relief =RIDGE)
          TitleFrame.pack(side=TOP)
          
          self.lblTitle = Label(TitleFrame,width=39,font=('Algerian',40,'bold'),bg="dark grey",fg="black",text="Library Management System",padx=12)
          self.lblTitle.grid()

          ButtonFrame = Frame(MainFrame,bd=20,width=1350,bg="dark grey",height =50,padx = 20)
          ButtonFrame.pack(side=BOTTOM)

          FrameDetail = Frame(MainFrame,bd =20, width =1350,height = 100,padx = 20,relief = RIDGE )
          FrameDetail.pack(side=BOTTOM)

          DataFrame=Frame(MainFrame,bd=20,width=1300,height=400,bg="dark grey",padx =20, relief =RIDGE)
          DataFrame.pack(side=BOTTOM)

          DataFrameLEFT =LabelFrame(DataFrame,bd = 10,width=800, height=300,bg="dark grey", padx = 20
                                    ,font=('arial',12,'bold'),text = "Library Memership Info:",)
          DataFrameLEFT.pack(side=LEFT)

          #book detail
          DataFrameRIGHT=LabelFrame(DataFrame,bd=10,width=450,height=300,bg="dark grey",padx=20          #list of all books in the library
                                    ,font=('arial',12,'bold'),text = "List of Books:")
          DataFrameRIGHT.pack(side=RIGHT)
          
          
#+++++++++++++++++++++++++++++++++++++++++++++++++++widgets+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++


      

          self.lblReference = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Book ID:",padx =2,pady =2)
          self.lblReference.grid(row=1,column=0,sticky=W)
          self.txtReference = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=Ref,width=25)
          self.txtReference.grid(row=1,column=1)
          
            #Book Title



          self.lblBookTitle = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Book Title:",padx =2,pady =2)
          self.lblBookTitle.grid(row=1,column=2,sticky=W)
          self.lblBookTitle = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=B_title,width=25)
          self.lblBookTitle.grid(row=1,column=3)

          #ppublication year


          self.lblyear = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Publ. Year:",padx =2,pady =2)
          self.lblyear.grid(row=2,column=0,sticky=W)
          self.lblyear = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=Year,width=25)
          self.lblyear.grid(row=2,column=1)
          #author

          self.lblAuthor = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Author:",padx =2,pady =2)
          self.lblAuthor.grid(row=2,column=2,sticky=W)
          self.lblAuthor = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=Author,width=25)
          self.lblAuthor.grid(row=2,column=3)
          #Book Type

          self.lblFirstname = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text ="Book Type:",padx =2,pady =2)
          self.lblFirstname.grid(row=3,column=0,sticky=W)
          self.lblFirstname = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=B_type,width=25)
          self.lblFirstname.grid(row=3,column=1)
          #Language
          
          self.lblDateLanguage = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Language :",padx =2,pady =2)
          self.lblDateLanguage.grid(row=3,column=2,sticky=W)
          self.lblDateLanguage = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=Language,width=25)
          self.lblDateLanguage.grid(row=3,column=3)
          #psize
          
          self.lblPsize = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "   Size:",padx =2,pady =2)
          self.lblPsize.grid(row=5,column=0,sticky=E)
          self.lblPsize = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=Psize,width=25)
          self.lblPsize.grid(row=5,column=1)



          #No of Page

          self.lblPostCode = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "No of Page",padx =2,pady =2)
          self.lblPostCode.grid(row=4,column=0,sticky=W)
          self.lblPostCode = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=Page,width=25)
          self.lblPostCode.grid(row=4,column=1)
          


          self.lblSellingPrice = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Selling Price:",padx =2,pady =2)
          self.lblSellingPrice.grid(row=4,column=2,sticky=W)
          self.lblSellingPrice = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=SellPrice,width=25)
          self.lblSellingPrice.grid(row=4,column=3)
#++++++++++++++++++++++++++++++++++Widgets++++++++++++++++++++++++++++++++++
          self.txtDisplay2=Text(DataFrameRIGHT,font=('arial',12,'bold'),width=32,bg="dark grey",height=13,padx=8,pady=20 )
          self.txtDisplay2.grid(row=0,column=2)

    
#+++++++++++++++++++++++++++++++++++Listbox++++++++++++++++++++++++++++++++
          
          
          ListOfBooks=[]                                                        #storage for booknames
          
          for row in sheet.iter_rows():                                         #iterate each row in column 1
                  ListOfBooks.append(row[1].value)                              #get all book titles from the excell database.
            
              
      
          def Selected_book(evt):                                               #get info for the selected book
                 sheet=File['Sheet1']                                           #importing sheet1 from the workbook

                 value=str(listbook.get(listbook.curselection()))
                 x = value                                                      #selected book
                 maxx_row = sheet.max_row                                       #maximum column
                 for i in range(1,maxx_row +1 ):
                     if x == (str(sheet.cell(row =i,column=2).value)):
                         B_title.set(sheet.cell(row=i,column=2).value)
                         Author.set(sheet.cell(row=i,column=3).value)
                         Page.set(sheet.cell(row=i,column=4).value)
                         Year.set(sheet.cell(row=i,column=5).value)
                         Psize.set(sheet.cell(row=i,column=7).value)
                         Language.set(sheet.cell(row=i,column=6).value)
                         Ref.set(sheet.cell(row=i,column=1).value)
                         SellPrice.set(sheet.cell(row=i,column=9).value)
                         B_type.set(sheet.cell(row=i,column=8).value)
                         
                     
                 Detail()
          
          
          listbook = Listbox(DataFrameRIGHT,width=20,height=12,font=('arial',12,'italic'))
          listbook.bind('<<ListboxSelect>>',Selected_book)                                              #gets the list of books that are retrived from our excell database
          listbook.grid(row=0,column=0,padx=8)
          
          scrollbar = Scrollbar(DataFrameRIGHT)
          scrollbar.grid(row=0,column=1,sticky='ns')
          scrollbar.config(command=listbook.yview)
    
          
          
          for items in ListOfBooks:                                                                     #appending each book into the listbook list                         
               listbook.insert(END,items)

            
#++++++++++++++++++++++++++++++++++Bottom textbox++++++++++++++++++++++++++++++++++
          self.lblLabel = Label(FrameDetail,font=('arial',14,'bold'),pady=8,              
                                text = " List Of Books Lent \n Title \t\t\t\t\t\t\t First name \t Last name \t Mob.No")
          self.lblLabel.grid(row=0,column = 0)
          
          self.txtDisplay=Text(FrameDetail,font=('arial',12,'bold'),width=140,height=4,padx=2,pady=4)
          self.txtDisplay.grid(row=1,column=0)
          
          sheet3=File['Sheet2']   
          maxx_row = sheet3.max_row
          for i in range(2,maxx_row +1):
              self.txtDisplay.insert(END, str(sheet3.cell(row=i,column=8).value) + "\t\t\t\t\t\t\t\t\t\t" + str(sheet3.cell(row=i,column=1).value)+ "\t\t\t"+str(sheet3.cell(row=i,column=2).value)+"\t\t"+str(sheet3.cell(row=i,column=4).value) +"\n")
              
         
           

#++++++++++++++++++++++++++++++++++Buttons++++++++++++++++++++++++++++++++++
          self.ButtonDis=Button(ButtonFrame,text = 'Add Book',font=('arial',12,'bold'),command=Add_Book,width=30 , bd=4)
          self.ButtonDis.grid(row=0, column=0)
          
          self.ButtonDel=Button(ButtonFrame,text = 'Log out',font=('arial',12,'bold'),command =Exit,width=30 , bd=4)
          self.ButtonDel.grid(row=0, column=1)

          self.ButtonReset=Button(ButtonFrame,text = 'Reset',font=('arial',12,'bold'),command=Reset,width=30 , bd=4)
          self.ButtonReset.grid(row=0, column=2)
          
          

class Book_Tracker_Stud:
     
     
     
     def __init__(self,root):
          self.root = root
          self.root.title("Library Management system")
          self.root.resizable(width =False, height=False)                     # to make the window size static
          self.root.configure(background='dark grey')
          Ref= StringVar()
          Fname= StringVar()
          Surname = StringVar()
          Address = StringVar()
    
          MobileNo = StringVar()
          B_title = StringVar()
          Author = StringVar()
          DateBorrowed = StringVar()
          DateDue = StringVar()
          SellPrice = StringVar()
          DateOverDue = StringVar()
          Language=StringVar()
          B_type=StringVar()
          Fine = StringVar()
          
          Page = StringVar()
          Year=StringVar()
          Psize=StringVar()
          
          os.chdir('C:\\Users\\OKSAS\\Desktop')                               #importing the excell file
          wb=openpyxl.load_workbook('managerdemo.xlsx')                       #excell file database
          sheet=wb['Sheet1']                                                  #importing sheet1 from the workbook
         

          def Del():
              Reset()
              self.txtDisplay2.delete("1.0",END)
          def Exit():
               Exit=tkinter.messagebox.askyesno("Library management system","Do you want to Log out ?")
               if Exit >0:
                    root.destroy()
                    return
          def Display():
              self.txtDesplay2.insert(END,int(Ref.get())+"\t"+B_title.get()+"\t"+Firstaname.get()+"\t"+Surname.get()+"\t\t"+ Address.get()+"\t"+"\t" + BookTitle.get()+"\t\t" + DateBorrowed.get() +"\t\t" + DaysOnLoan.get() + "\n")
          def Detail():                                                                                                #displaying the details of the selected book 
              self.txtDisplay2.config(state=NORMAL)                                                                     
              
              self.txtDisplay2.delete("1.0",END)                                                                       #clearing the text entries 
              self.txtDisplay2.insert(END,'Title: \t\t' + B_title.get() + "\n")
              self.txtDisplay2.insert(END,'Author: \t\t' + Author.get() + "\n" )
              self.txtDisplay2.insert(END,'No of page: \t\t' + Page.get() + "\n" )
                  
              
              self.txtDisplay2.insert(END,'Language :\t\t' + Language.get()+ "\n" )
              self.txtDisplay2.insert(END,'Size: \t\t' + Psize.get() + "\n" )
              self.txtDisplay2.insert(END,'Year: \t\t' + Year.get() + "\n" )
              self.txtDisplay2.insert(END,'Book ID: \t\t' + Ref.get() + "\n" )
              self.txtDisplay2.insert(END,'Type: \t\t' + B_type.get() + "\n" )
         
             
              
              
              
              self.txtDisplay2.config(state=DISABLED)                                                                 #Preventing  data temeper

          listInfo2=[]
 #++++++++++++++++++++++++++++++++++++++++++++++++bootom detail++++++++++++++++++++++++++++++++++
          def rent():
              
              self.txtDisplay.config(state=NORMAL)
                                                                                                                     #clearing the text entries 
              self.txtDisplay.insert(END,B_title.get() + "\t\t\t" )
              self.txtDisplay.insert(END,Author.get() + "\t\t")
              self.txtDisplay.insert(END, Page.get() + "\t\t")
              self.txtDisplay.insert(END,Language.get()+ "\t\t" )
              self.txtDisplay.insert(END,Year.get() + "\t\t" )
              self.txtDisplay.insert(END, DateBorrowed.get() + "\t\t" )
              self.txtDisplay.insert(END, DateDue.get() + "\t\t" )
              self.txtDisplay.insert(END, SellPrice.get() + "\t\t" )
            
              
              self.txtDisplay2.config(state=DISABLED)                                                                 #Preventing  data temeper
              listInfo=[Fname.get(),Surname.get(),Address.get(),int(MobileNo.get()),int(Ref.get()),DateBorrowed.get(),DateDue.get(),B_title.get(),Author.get()]
              
              listInfo2.append(listInfo)
              final(listInfo2 )
              
            
          def final(listt):                                                                                            #Borrower information
               sheet2=wb['Sheet2']
               maxx_row=sheet2.max_row
               maxx_column= sheet2.max_column
               for j in range(len(listt)):    
                   for i in range(1,maxx_column+1):
                       sheet2.cell(row=maxx_row + 1,column=i).value=listt[j][i-1]
                       wb.save('managerdemo.xlsx')
             

             
             

             
          def Reset():

               Ref.set("")
               Fname.set("")
               Surname.set("")
               Address.set("")
               MobileNo.set("")
               B_title.set("")
               Author.set("")
               DateBorrowed.set("")
               DateDue.set("")
               SellPrice.set("")
               Fine.set("")
               Year.set("")
               Language.set("")
               Page.set("")
               
           
               self.txtDisplay.delete("1.0",END)
#main frame
          MainFrame = Frame(self.root)
          MainFrame.grid()

          TitleFrame = Frame(MainFrame,width = 1350,bg="dark grey",padx = 20,bd =20,relief ='groove')
          TitleFrame.pack(side=TOP)
          
          self.lblTitle = Label(TitleFrame,width=39,font=('Algerian',40,'bold'),bg="dark grey",fg="Black",text="Library Management System",padx=12)
          self.lblTitle.grid()

          ButtonFrame = Frame(MainFrame,bd=20,width=1350,bg="dark grey",height =50,padx = 20, relief = RIDGE)
          ButtonFrame.pack(side=BOTTOM)

          FrameDetail = Frame(MainFrame,bd =20, width =1350,height = 100,padx = 20,relief = RIDGE )
          FrameDetail.pack(side=BOTTOM)

          DataFrame=Frame(MainFrame,bd=20,width=1300,height=400,bg="dark grey",padx =20, relief =RIDGE)
          DataFrame.pack(side=BOTTOM)

          DataFrameLEFT =LabelFrame(DataFrame,bd = 10,width=800, height=300,bg="dark grey", padx = 20,relief=RIDGE
                                    ,font=('arial',12,'bold'))
          DataFrameLEFT.pack(side=LEFT)

          #book detail
          DataFrameRIGHT=LabelFrame(DataFrame,bd=10,width=450,height=300,bg="dark grey",padx=20,relief =RIDGE
                                    ,font=('arial',12,'bold'),text = "Book Details:")
          DataFrameRIGHT.pack(side=RIGHT)
          
          
#widgets

          self.lblReference = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Book ID:",padx =2,pady =2)
          self.lblReference.grid(row=1,column=0,sticky=W)
          self.txtReference = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=Ref,width=25)
          self.txtReference.grid(row=1,column=1)
          
          #

          self.lblBookTitle = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Book Title:",padx =2,pady =2)
          self.lblBookTitle.grid(row=2,column=0,sticky=W)
          self.lblBookTitle = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=B_title,width=25)
          self.lblBookTitle.grid(row=2,column=1)

    
          #author

          self.lblAuthor = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Author:",padx =2,pady =2)
          self.lblAuthor.grid(row=2,column=2,sticky=W)
          self.lblAuthor = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=Author,width=25)
          self.lblAuthor.grid(row=2,column=3)
          #first name

          self.lblFirstname = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "First name:",padx =2,pady =2)
          self.lblFirstname.grid(row=3,column=0,sticky=W)
          self.lblFirstname = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=Fname,width=25)
          self.lblFirstname.grid(row=3,column=1)
          #date borrowed
          
          self.lblDateBorrowed = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Date Borrowed :",padx =2,pady =2)
          self.lblDateBorrowed.grid(row=3,column=2,sticky=W)
          self.lblDateBorrowed = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=DateBorrowed,width=25)
          self.lblDateBorrowed.grid(row=3,column=3)
          #surname

          self.lblSurname = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Surname :",padx =2,pady =2)
          self.lblSurname.grid(row=4,column=0,sticky=W)
          self.lblSurname = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=Surname,width=25)
          self.lblSurname.grid(row=4,column=1)
          #DateDue

          self.lblDateDue = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Date Due :",padx =2,pady =2)
          self.lblDateDue.grid(row=4,column=2,sticky=W)
          self.lblDateDue = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=DateDue,width=25)
          self.lblDateDue.grid(row=4,column=3)
          #Address

          self.lblAddress1 = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Address:",padx =2,pady =2)
          self.lblAddress1.grid(row=5,column=0,sticky=W)
          self.lblAddress1 = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=Address,width=25)
          self.lblAddress1.grid(row=5,column=1)
          
        

          #Late return Fine

          self.lblLateReturnFine = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Late Return Fine:",padx =2,pady =2)
          self.lblLateReturnFine.grid(row=6,column=2,sticky=W)
          self.txtLateReturnFine = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=Fine,width=25)
          self.txtLateReturnFine.grid(row=6,column=3)

          #No of Page

          self.lblPostCode = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "No of Page",padx =2,pady =2)
          self.lblPostCode.grid(row=1,column=2,sticky=W)
          self.lblPostCode = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=Page,width=25)
          self.lblPostCode.grid(row=1,column=3)
          
          #Mobile Number

          
          self.lblMobileNo = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Mobile no:",padx =2,pady =2)
          self.lblMobileNo.grid(row=5,column=2,sticky=W)
          self.lblMobileNo = Entry(DataFrameLEFT,font =('arial',12,'bold'),width=25,textvariable=MobileNo)
          self.lblMobileNo.grid(row=5,column=3)
          #Selling Price

          self.lblSellingPrice = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Selling Price:",padx =2,pady =2)
          self.lblSellingPrice.grid(row=6,column=0,sticky=W)
          self.lblSellingPrice = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=SellPrice,width=25)
          self.lblSellingPrice.grid(row=6,column=1)
#++++++++++++++++++++++++++++++++++Widget++++++++++++++++++++++++++++++++++
          self.txtDisplay2=Text(DataFrameRIGHT,font=('arial',12,'bold'),width=32,bg="grey",height=13,padx=8,pady=20 )
          self.txtDisplay2.grid(row=0,column=2)

          
          self.lblLanguage = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Language:",padx =2,pady =2)
          self.lblLanguage = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=Language,width=25)
         
           
          self.lblYear = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Year:",padx =2,pady =2)
          self.lblYear = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=Year,width=25)

          self.lblPsize = Label(DataFrameLEFT,font =('arial',12,'bold'),bg="dark grey",text = "Size:",padx =2,pady =2)
          self.lblPsize = Entry(DataFrameLEFT,font =('arial',12,'bold'),textvariable=Psize,width=25)
         

          
          
#+++++++++++++++++++++++++++++++++++Listbox++++++++++++++++++++++++++++++++
          
          
          ListOfBooks=[]                                                      #storage for booknames
          
          for row in sheet.iter_rows():                                       #iterate each row in column 1
                  ListOfBooks.append(row[1].value)                          #get all book titles from the excell database.
              
      
          def Selected_book(evt):                                             #get info for the selected book
                 value=str(listbook.get(listbook.curselection()))
                 x = value
                 maxx_row = sheet.max_row                                       #maximum column
                 for i in range(1,maxx_row + 1):
                     import datetime
                     if x == (str(sheet.cell(row =i,column=2).value)):
                         B_title.set(sheet.cell(row=i,column=2).value)
                         Author.set(sheet.cell(row=i,column=3).value)
                         Page.set(sheet.cell(row=i,column=4).value)
                         Year.set(sheet.cell(row=i,column=5).value)
                         today=datetime.date.today()                       #the date the user borrowed the book
                         DateBorrowed.set(today)
                         Lday=datetime.timedelta(days=10)                 
                         Rday = today + Lday                               # return date
                         DateDue.set(Rday)                                 #due date
                         Psize.set(sheet.cell(row=i,column=7).value)
                         Language.set(sheet.cell(row=i,column=6).value)
                         Ref.set(sheet.cell(row=i,column=1).value)
                         SellPrice.set(sheet.cell(row=i,column=9).value)
                         B_type.set(sheet.cell(row=i,column=8).value)
                         Fine.set(sheet.cell(row=i,column=10).value)
                     
                         
                 Detail()
          
          
          listbook = Listbox(DataFrameRIGHT,width=20,height=12,font=('arial',12,'italic'))
          listbook.bind('<<ListboxSelect>>',Selected_book)
          listbook.grid(row=0,column=0,padx=8)
          
          scrollbar = Scrollbar(DataFrameRIGHT)
          scrollbar.grid(row=0,column=1,sticky='ns')
          scrollbar.config(command=listbook.yview)
          scrollbar.config(command=self.txtDisplay2.xview)
          
          
          for items in ListOfBooks:                                                                         #appending books to the listbook list
               listbook.insert(END,items)

            
#++++++++++++++++++++++++++++++++++Bottom Lable++++++++++++++++++++++++++++++++++
          self.lblLabel = Label(FrameDetail,font=('arial',10,'bold'),pady=8,
                                text = "Title \t\t\t\t Author\t\t size \t\t Language \t\t year \t\t Date Borrowed\t\t\t\t Days Due \t\t ")
          self.lblLabel.grid(row=0,column = 0)
          self.txtDisplay=Text(FrameDetail,font=('arial',12,'bold'),width=140,height=4,padx=2,pady=4)
          self.txtDisplay.grid(row=1,column=0)
               

#++++++++++++++++++++++++++++++++++Buttons++++++++++++++++++++++++++++++++++
          self.ButtonDis=Button(ButtonFrame,text = 'Rent',font=('arial',12,'bold'),command=rent,width=30 , bd=4)
          self.ButtonDis.grid(row=0, column=0)

          self.ButtonDel=Button(ButtonFrame,text = 'Log out',font=('arial',12,'bold'),command =Exit,width=30 , bd=4)
          self.ButtonDel.grid(row=0, column=1)

          self.ButtonReset=Button(ButtonFrame,text = 'Reset',font=('arial',12,'bold'),command=Reset,width=30 , bd=4)
          self.ButtonReset.grid(row=0, column=2)
main()
